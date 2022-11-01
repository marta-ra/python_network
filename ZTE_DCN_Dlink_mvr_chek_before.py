from netmiko import ConnectHandler
import csv
import re
import os
from openpyxl import load_workbook
import time
from datetime import datetime

LOGIN = ''
PASSWORD = ''
vlan_uplik = ''
multi_filter = ''
number_R = str(input('Input number of R: '))


def writer_log_file(ip_host, command, output):
    global number_R
    name_file = 'MVR_check_logfile.csv'
    time_command = str(datetime.now())
    try:
        with open(name_file, 'a', newline='') as csvfile:
            spamwriter = csv.writer(csvfile, delimiter=';')
            spamwriter.writerow([time_command, 'R' + number_R, ip_host, str(command), str(output)])
            return None
    except:
        return 'File error'


def writer_ZTE_DCN_Dlink_check_file(*args):
    global number_R
    global model
    name_file = f'{model}_R{number_R}_check.csv'
    row_write = []
    time_command = str(datetime.now())
    row_write.append(time_command)
    for value in args:
        row_write.append(value)
    try:
        with open(name_file, 'a', newline='') as csvfile:
            spamwriter = csv.writer(csvfile, delimiter=';')
            spamwriter.writerow(row_write)
            return None
    except:
        return 'File error'

def result(success, search_for):
    print(success)
    writer_ZTE_DCN_Dlink_check_file(new_mgmt_ip, search_for, success)


path = r'D:\XXX'
workbook = load_workbook(os.path.join(path, "IP.xlsx"))

# Choose active sheet:
for sheet in workbook:
    sheet_name = 'R' + number_R
    if sheet.title == sheet_name:
        workbook[sheet.title].views.sheetView[0].tabSelected = True
    else:
        workbook[sheet.title].views.sheetView[0].tabSelected = False
workbook.active = workbook[sheet_name]
worksheet = workbook.active
max_row_not_empty = worksheet.max_row


for rownum in range(1,max_row_not_empty):
    model = str(worksheet.cell(row=rownum, column=8).value)
    print(model)
    print(rownum)
    new_mgmt_ip = str(worksheet.cell(row=rownum, column=5).value)
    customer_vlan = str(worksheet.cell(row=rownum, column=10).value)
        
    print('\n+++++++++++++++++++++++++++++++++')

    if 'ZTE5260' in model:
        device = {
            "device_type": "cisco_ios_telnet",
            "ip": new_mgmt_ip,
            "username": LOGIN,
            "password": PASSWORD,
                }
        print(f'connet to switch {new_mgmt_ip} .....')

        with ConnectHandler(**device) as ssh:
            ssh.enable()
            print(f'conneted to switch {new_mgmt_ip}')
            output =  ssh.send_command('show version', expect_string=r"#")
            writer_log_file(new_mgmt_ip, 'show version', output)
            print(output)
            ports_count = int((re.search(r'\d\d',(re.search(r'-\d\dTD-H Software', output)).group(0))).group(0))
            print(ports_count)
            output =  ssh.send_command(f'show vlan id {customer_vlan}', expect_string=r"#")
            writer_log_file(new_mgmt_ip, f'show vlan id {customer_vlan}', output)
            print(output)
            port_ethernet_syntax = (re.search(r'gei-\d/\d/\d/', output)).group(0)
            print(port_ethernet_syntax)
            port_uplink_syntax = (re.search(r'xgei-\d/\d/\d/', output)).group(0)
            print(port_uplink_syntax)
            for port in range(1, ports_count-3):
                print(f'show running-config-interface {port_ethernet_syntax}{port}')
                output =  ssh.send_command(f'show running-config-interface {port_ethernet_syntax}{port}', expect_string=r"#")
                writer_log_file(new_mgmt_ip, f'show running-config-interface {port_ethernet_syntax}{port}', output)
                print(output)
                try:
                    hybrid_and_vlan = (re.search(fr'switchport hybrid vlan.*{customer_vlan}.*untag', output)).group(0)
                    print(hybrid_and_vlan)
 
                    result = 'OK'
                    writer_ZTE_DCN_Dlink_check_file(new_mgmt_ip, port_ethernet_syntax + str(port), hybrid_and_vlan, result)
                except AttributeError as error:
                    print(error)

                    result = 'Warning'
                    writer_ZTE_DCN_Dlink_check_file(new_mgmt_ip, port_ethernet_syntax + str(port), output, result)
            
            for port in range(ports_count-3, ports_count+1):
                print(f'show running-config-interface {port_uplink_syntax}{port}')
                output =  ssh.send_command(f'show running-config-interface {port_uplink_syntax}{port}', expect_string=r"#")
                writer_log_file(new_mgmt_ip, f'show running-config-interface {port_uplink_syntax}{port}', output)
                print(output)
                try:
                    hybrid_and_vlan = (re.search(fr'switchport trunk vlan.*{customer_vlan}.*', output)).group(0)
                    print(hybrid_and_vlan)
      
                    result = 'OK'
                    writer_ZTE_DCN_Dlink_check_file(new_mgmt_ip, port_uplink_syntax + str(port), hybrid_and_vlan, result)
                except AttributeError as error:
                    print(error)

                    result = 'Warning'
                    writer_ZTE_DCN_Dlink_check_file(new_mgmt_ip, port_uplink_syntax + str(port), output, result)
            

        ConnectHandler(**device).disconnect()
        print('\n-------------------------------------------------')

    elif 'ZTE5928' in model:
        device = {
            "device_type": "cisco_ios_telnet",
            "ip": new_mgmt_ip,
            "username": LOGIN,
            "password": PASSWORD,
                }
        print(f'connet to switch {new_mgmt_ip} .....')

        with ConnectHandler(**device) as ssh:
            ssh.enable()
            print(f'conneted to switch {new_mgmt_ip}')
            port_ethernet_syntax = 'gei-0/1/1/'
            port_uplink_syntax = 'gei-0/1/2/'
            ports_count = 28
            for port in range(1, ports_count-3):
                print(f'show running-config-interface {port_ethernet_syntax}{port}')
                output =  ssh.send_command(f'show running-config-interface {port_ethernet_syntax}{port}', expect_string=r"#")
                writer_log_file(new_mgmt_ip, f'show running-config-interface {port_ethernet_syntax}{port}', output)
                print(output)
                try:
                    hybrid_and_vlan = (re.search(fr'switchport hybrid vlan.*{customer_vlan}.*untag', output)).group(0)
                    print(hybrid_and_vlan)
 
                    result = 'OK'
                    writer_ZTE_DCN_Dlink_check_file(new_mgmt_ip, port_ethernet_syntax + str(port), hybrid_and_vlan, result)
                except AttributeError as error:
                    print(error)

                    result = 'Warning'
                    writer_ZTE_DCN_Dlink_check_file(new_mgmt_ip, port_ethernet_syntax + str(port), output, result)
            
            for port in range(1, 5):
                print(f'show running-config-interface {port_uplink_syntax}{port}')
                output =  ssh.send_command(f'show running-config-interface {port_uplink_syntax}{port}', expect_string=r"#")
                writer_log_file(new_mgmt_ip, f'show running-config-interface {port_uplink_syntax}{port}', output)
                print(output)
                try:
                    hybrid_and_vlan = (re.search(fr'switchport trunk vlan.*{customer_vlan}.*', output)).group(0)
                    print(hybrid_and_vlan)
      
                    result = 'OK'
                    writer_ZTE_DCN_Dlink_check_file(new_mgmt_ip, port_uplink_syntax + str(port), hybrid_and_vlan, result)
                except AttributeError as error:
                    print(error)

                    result = 'Warning'
                    writer_ZTE_DCN_Dlink_check_file(new_mgmt_ip, port_uplink_syntax + str(port), output, result)
            

        ConnectHandler(**device).disconnect()
        print('-------------------------------------------------')



    elif 'DCN' in model or 'QSW' in model:
        device = {
            "device_type": "cisco_ios_telnet",
            "ip": new_mgmt_ip,
            "username": LOGIN,
            "password": PASSWORD,
            "fast_cli": False,
                }
        print(f'connet to switch {new_mgmt_ip} .....')


        with ConnectHandler(**device) as ssh:
            ssh.enable()
            print(f'conneted to switch {new_mgmt_ip}')
            time.sleep(2)
            output =  ssh.send_command(f'show vlan id {vlan_uplik }', delay_factor = 10)
            if re.search(r'enable', output):
                output =  ssh.send_command(f'show vlan id {vlan_uplik }', delay_factor = 10)
            print(output)
            port_find = re.findall(r'Ethernet\S*\(.\)', output)
            list_for_calc_ethernet_ports = [] # список аплинков
            print(port_find)
            for value in port_find:
                trunk = (((re.search(r'\(.\)', value)).group(0)).replace('(', '')).replace(')', '')
                print(trunk)
                port = value.replace((re.search(r'\(.\)', value)).group(0), '')
                print(port)
                port_num = ((re.search(r'/\d*$', port)).group(0)).replace('/', '')  # для записи номеров аплинков
                print(port_num)
                list_for_calc_ethernet_ports.append(port_num)
                print(list_for_calc_ethernet_ports)
        
                if trunk == 'T':
                    writer_ZTE_DCN_Dlink_check_file(new_mgmt_ip, port, trunk, vlan_uplik, 'OK')
                else:
                    writer_ZTE_DCN_Dlink_check_file(new_mgmt_ip, port, trunk, vlan_uplik , 'Warning')

            port_syntax = port.replace(port_num, '')
            time.sleep(2)
            output =  ssh.send_command('show ip multicast destination-control', delay_factor = 10)
            if len(output) == 0:
                output =  ssh.send_command('show ip multicast destination-control', delay_factor = 10)
            print(output)
            
            first_uplink = min(list_for_calc_ethernet_ports)
            print(first_uplink)

            for port in range(1, int(first_uplink)):
                ethernet_port = port_syntax + str(port)
                print(ethernet_port)
                try:
                    multicast_port = (re.search(fr'multicast destination-control access-group \d\d\d\d used on interface {ethernet_port}\b', output)).group(0)
                    result = 'OK'
                    print(result)
                    writer_ZTE_DCN_Dlink_check_file(new_mgmt_ip, ethernet_port, 'ethernet', multicast_port, result)
                except AttributeError as error:
                    print(error)
                    result = 'Warning'
                    print(result)
                    writer_ZTE_DCN_Dlink_check_file(new_mgmt_ip, ethernet_port, 'ethernet', error, result)

        ConnectHandler(**device).disconnect()
        print('-------------------------------------------------')

    
    elif 'Dlink' in model:
    
        find_uplink = f'config vlan {vlan_uplik} add tagged 25-28'
        device = {
            "device_type": "cisco_ios_telnet",
            "ip": new_mgmt_ip,
            "username": LOGIN,
            "password": PASSWORD,
            "fast_cli": False,
                }
        print(f'connet to switch {new_mgmt_ip} .....')

        with ConnectHandler(**device) as ssh:
            ssh.enable()
            print(f'conneted to switch {new_mgmt_ip}')
            output =  ssh.send_command('show config current_config', delay_factor = 50, expect_string=r"#")
            # print(output)
            if re.search(multi_filter, output):
                result('OK', 'multi_filter')
            else:
                result('ERROR', 'multi_filter')
            if re.search(find_uplink, output):
                result('OK', 'uplinks')
            else:
                result('ERROR', 'uplinks')


        ConnectHandler(**device).disconnect()
        print('\n-------------------------------------------------')
