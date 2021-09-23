from netmiko import ConnectHandler
import csv
import re
import os
from openpyxl import load_workbook
import time
from datetime import datetime


LOGIN = 'login'
PASSWORD = 'password'
model_ZTE = 'ZTE'


def writer_log_file(ip_host, command, output):
    name_file = 'ZTE_check_result_logfile.csv'
    time_command = str(datetime.now())
    try:
        with open(name_file, 'a', newline='') as csvfile:
            spamwriter = csv.writer(csvfile, delimiter=';')
            spamwriter.writerow([time_command, ip_host, str(command), str(output)])
            return None
    except:
        return 'File error'

def writer_check_port_file(ip_host, port, output, result):
    name_file = 'ZTE_check_result.csv'
    time_command = str(datetime.now())
    try:
        with open(name_file, 'a', newline='') as csvfile:
            spamwriter = csv.writer(csvfile, delimiter=';')
            spamwriter.writerow([time_command, ip_host, port, output, result])
            return None
    except:
        return 'File error'


path = r'D:\XXX\XXX'
number_R = str(input('Input number of R: '))
workbook = load_workbook(os.path.join(path, "IP_plan.xlsx"))

# Choose active sheet:
for sheet in workbook:
    sheet_name = 'R' + number_R
    if sheet.title == sheet_name:
        workbook[sheet.title].views.sheetView[0].tabSelected = True
    else:
        workbook[sheet.title].views.sheetView[0].tabSelected = False
workbook.active = workbook[sheet_name]


for rownum in range(64,65):
    print(rownum)
    model = str(workbook.active.cell(row=rownum, column=8).value)
    new_hostname = str(workbook.active.cell(row=rownum, column=2).value)
    new_mgmt_ip = str(workbook.active.cell(row=rownum, column=5).value)
    mgmt_vlan = str(workbook.active.cell(row=rownum, column=3).value)
    mask = str(workbook.active.cell(row=rownum, column=6).value)
    mgmt_gw = str(workbook.active.cell(row=rownum, column=7).value)
    mvr_vlan = str(workbook.active.cell(row=rownum, column=9).value)
    customer_vlan = str(workbook.active.cell(row=rownum, column=10).value)

    if model_ZTE in model:
        device = {
            "device_type": "cisco_ios_telnet",
            "ip": new_mgmt_ip,
            "username": LOGIN,
            "password": PASSWORD,
                }
        print(f'connet to switch {new_mgmt_ip} .....')

        with ConnectHandler(**device) as ssh:
            ssh.enable()
            print(f'conneted to switch {new_mgmt_ip}')
            output =  ssh.send_command('show version', expect_string=r"#")
            print(output)
            writer_log_file(new_mgmt_ip, 'show version', output)
            # search for the number of ports:
            ports_count = int((re.search(r'\d\d',(re.search(r'-\d\dTD-H Software', output)).group(0))).group(0))
            print(ports_count)
            output =  ssh.send_command(f'show vlan id {customer_vlan}', expect_string=r"#")
            writer_log_file(new_mgmt_ip, 'show vlan id {customer_vlan}', output)
            print(output)
            # search for spelling ports:
            port_client_syntax = (re.search(r'gei-\d/\d/\d/', output)).group(0)
            print(port_client_syntax)
            port_uplink_syntax = (re.search(r'xgei-\d/\d/\d/', output)).group(0)
            print(port_uplink_syntax)
            # checking a specific setting on a port:
            for port in range(1, ports_count-4):
                output =  ssh.send_command(f'show running-config-interface {port_client_syntax}{port}', expect_string=r"#")
                writer_log_file(new_mgmt_ip, f'show running-config-interface {port_client_syntax}{port}', output)
                print(output)
                try:
                    hybrid_and_vlan = (re.search(fr'switchport hybrid vlan.*{customer_vlan}.*untag', output)).group(0)
                    print(hybrid_and_vlan)
                    result = 'OK'
                    writer_check_port_file(new_mgmt_ip, port_client_syntax + str(port), hybrid_and_vlan, result)
                except AttributeError as error:
                    print(error)
                    result = 'Warning'
                    writer_check_port_file(new_mgmt_ip, port_client_syntax + str(port), output, result)

        ConnectHandler(**device).disconnect()
        print('-------------------------------------------------')
