# pars interfase statistic for find ports(uplinks) information

from netmiko import ConnectHandler
import csv
import re
import os
from openpyxl import load_workbook


LOGIN = 'NA'
PASSWORD = '8h/;E#Aai^NVRQ'

port_0 = '0/0/'
port_1 = '1/'
port_1_0 = '1/0/'

numport_list_28 = ['24','25','26','27','28']
numport_list_52 = ['48','49','50','51','52']


def writer_log_file(number_distr, name_host, host, command, port, port_stat, protocol_stat, alias_name):
    
    name_file = 'log_dcn_check' + str(number_distr) +'.csv'
    try:
        with open(name_file, 'a', newline='') as csvfile:
            spamwriter = csv.writer(csvfile, delimiter=';')
            spamwriter.writerow([name_host, host, command, 'ethernet ' + port, port_stat, protocol_stat, alias_name])
            return None
    except:
        return 'File error'

def find_port(output,numport):
    
    output = ssh.send_command('show interface ethernet ' + port_0 + numport)
    port_right_or_not = output.find('interface error!')
    if port_right_or_not == -1:
        port = port_0
        return port
    else:
        output = ssh.send_command('show interface ethernet ' + port_1 + numport)
        port_right_or_not_1 = output.find('interface error!')
        if port_right_or_not_1 == -1:
            port = port_1
            return port
        else:
            port = port_1_0
            output = ssh.send_command('show interface ethernet ' + port_1_0 + numport)
            port_right_or_not_1 = output.find('interface error!')
            if port_right_or_not_1 == -1:
                return port
            else:
                return 'interface error!'

def find_port_list():
    numport = numport_list_52[3]
    output_show =  ssh.send_command('show interface ethernet ' + port_0 + numport)
    lexis_port = find_port(output_show, numport)
    if lexis_port == 'interface error!':
        numport = numport_list_28[3]
        lexis_port = find_port(output_show, numport)
        return numport_list_28, lexis_port
    else:
        return numport_list_52, lexis_port



path = r'X\X\X'
number_distr = input('Input number of district: ')
tg_xlsx = load_workbook(os.path.join(path, "nodes_DCN" +str(number_distr) +".xlsx"), read_only=True)
tg_sheet = tg_xlsx.active


for rownum in range(2,5):
    print(rownum)
    name_switch = str(tg_sheet.cell(row=rownum, column=1).value)
    ip_switch = str(tg_sheet.cell(row=rownum, column=2).value)
    print(name_switch)
    print(ip_switch)

    device = {
        "device_type": "cisco_ios_telnet",
        "ip": ip_switch,
        "username": LOGIN,
        "password": PASSWORD,
            }
    print(f'connet to {ip_switch}')

    with ConnectHandler(**device) as ssh:
        ssh.enable()

        print('find port...')
        port_data = find_port_list()
        print(port_data)
        lexis_port = port_data[1]
        numport_list = port_data[0]
        print(lexis_port)
        print(numport_list)
        for numport in numport_list:
            print('show interface ethernet ' + lexis_port + numport)
            output_show =  ssh.send_command('show interface ethernet ' + lexis_port + numport)
            # print('output: '+ output_show)

            port_stat = (re.search(r'Ethernet' + lexis_port + numport + r' is \D*,', output_show)).group(0)
            protocol_stat = (re.search(r' line protocol is \D*\b', output_show)).group(0)
            alias_name = (re.search(r', alias name is .*,', output_show)).group(0)
            
            port_stat = port_stat.replace(r'Ethernet' + lexis_port + numport + r' is ', '')
            port_stat = port_stat.replace(',', '')
            protocol_stat = protocol_stat.replace(' line protocol is ', '')
            protocol_stat = protocol_stat.replace('\n', '')
            alias_name = alias_name.replace(', alias name is ', '')
            alias_name = alias_name.replace(',', '')
            print('---------------')
            print(port_stat)
            print(protocol_stat)
            print(alias_name)
            writer_log_file(number_distr, name_switch, ip_switch, 'show interface ethernet ' + lexis_port + numport, lexis_port + numport, port_stat, protocol_stat, alias_name)


        ConnectHandler(**device).disconnect()
        print('-------------------------------------------------')



